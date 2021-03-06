from load import *
from portal.models import Permiso
from base.models import Institucion, Raw
from usuarios.models import User

# Raw.objects.filter(anio=2021).delete()

import xlsxwriter


from openpyxl import load_workbook
wb = load_workbook(filename='../desarrollo/Resultados2021/activa.xlsx')
wsi = wb[wb.sheetnames[0]]

CODIGOS = {
    'Subsecretaría de Bienes Nacionales': '1401',
    'Caja de Previsión de la Defensa Nacional': '1513',
    'Dirección General de Crédito Prendario': '1504',
    'Dirección de Previsión de Carabineros de Chile': '1514',
    'Dirección del Trabajo': '1502',
    'INSTITUTO DE PREVISION SOCIAL': '1509',
    'Instituto de Seguridad Laboral': '1510',
    'Superintendencia de Seguridad Social': '1506',
    'Servicio Nacional de Capacitación y Empleo': '1505',
    'Superintendencia de Pensiones': '1507',
    'Superintendencia de Electricidad y Combustibles': '2404',

}

codigo_dipres_anterior = None
raws = []
index = 0
instituciones_ids = []
for row in wsi.iter_rows():
    if row[0].value is None:
        continue

    if row[0].value in ['CODIGOS_INSTITUCIONES']:
        headers = [r.value for r in row]
        continue

    cod_institucion = str(row[0].value)
    if len(cod_institucion) == 3:
        cod_institucion = '0' + cod_institucion

    # cod_institucion = CODIGOS[cod_institucion]

    # institucion = Institucion.objects.get(datos__codigo_dipres=cod_institucion)

    datos = {}
    for i, header in enumerate(headers):
        datos[header] = row[i].value

    if codigo_dipres_anterior is None or codigo_dipres_anterior != cod_institucion:
        institucion = Institucion.objects.get(codigo_dipres=cod_institucion)
        codigo_dipres_anterior = cod_institucion
        print(institucion.nombre)

    datos["file"] = "activa.xlsx"
    datos["index"] = index + 1
    raws.append(Raw(institucion=institucion, anio=2021, datos=datos))
    instituciones_ids.append(institucion.id)
    index += 1

Raw.objects.filter(anio=2021, institucion_id__in=instituciones_ids).delete()
resp = Raw.objects.bulk_create(raws)

instituciones_ids = list(set(instituciones_ids))

completos = []
for institucionId in instituciones_ids:
    institucion = Institucion.objects.get(pk=institucionId)
    workbook = xlsxwriter.Workbook('../2021/{}.xlsx'.format(institucion.nombre))
    worksheet = workbook.add_worksheet()
    row = 0
    for raw in raws:
        if raw.institucion.id != institucion.id:
            continue
        if row == 0:
            for col, columna in enumerate(raw.datos.keys()):
                a = worksheet.write(row, col, columna)

        for col, columna in enumerate(raw.datos.keys()):
            b = worksheet.write(row + 1, col, raw.datos.get(columna))

        row += 1

    workbook.close()


7 / 0
codigos = []
for raw in raws:
    if raw.codigo_cf not in codigos:
        codigos.append(raw.codigo_cf)


def limpiar(valor):
    if valor is None:
        return 0
    if valor.find("NO LEER") > 0:
        return None
    return int(valor.split(" ")[0])


def limpiar_pond(valor):
    if valor == '-':
        return 0
    return valor


completos = []
for codigo in codigos:
    experiencia_positivo = 0
    experiencia_negativo = 0
    experiencia_total = 0
    eval_inst_positivo = 0
    eval_inst_negativo = 0
    eval_inst_total = 0
    cantidad = 0
    workbook = xlsxwriter.Workbook('../2021/{}.xlsx'.format(codigo))
    worksheet = workbook.add_worksheet()
    row = 0
    for raw in raws:
        if raw.codigo_cf != codigo:
            continue
        if row == 0:
            for col, columna in enumerate(raw.datos.keys()):
                a = worksheet.write(row, col, columna)

        for col, columna in enumerate(raw.datos.keys()):
            b = worksheet.write(row + 1, col, raw.datos.get(columna))

        row += 1

        experiencia_positivo += limpiar_pond(raw.datos.get("PONDERADOR"))if limpiar(raw.datos.get("PEX05")) in [6, 7] else 0
        experiencia_negativo += limpiar_pond(raw.datos.get("PONDERADOR"))if limpiar(raw.datos.get("PEX05")) in [1, 2, 3, 4] else 0
        experiencia_total += limpiar_pond(raw.datos.get("PONDERADOR"))if limpiar(raw.datos.get("PEX05")) in [1, 2, 3, 4, 5, 6, 7] else 0
        eval_inst_positivo += limpiar_pond(raw.datos.get("PONDERADOR"))if limpiar(raw.datos.get("PI02")) in [6, 7] else 0
        eval_inst_negativo += limpiar_pond(raw.datos.get("PONDERADOR"))if limpiar(raw.datos.get("PI02")) in [1, 2, 3, 4] else 0
        eval_inst_total += limpiar_pond(raw.datos.get("PONDERADOR"))if limpiar(raw.datos.get("PI02")) in [1, 2, 3, 4, 5, 6, 7] else 0
        cantidad += 1

    institucion = Institucion.objects.get(datos__codigo_dipres=codigo)
    data = {
        'institucion': institucion.nombre,
        'cod_institucion': codigo,
        'experiencia_positivo': experiencia_positivo,
        'experiencia_negativo': experiencia_negativo,
        'experiencia_total': experiencia_total,
        'eval_inst_positivo': eval_inst_positivo,
        'eval_inst_negativo': eval_inst_negativo,
        'eval_inst_total': eval_inst_total,
        'cantidad': cantidad

    }
    workbook.close()
    with open('../2021/{}.json'.format(codigo), 'w') as outfile:
        json.dump(data, outfile)

    completos.append(data)


with open('../2021/activa.json', 'w') as outfile:
    json.dump(completos, outfile)
